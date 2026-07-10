#!/usr/bin/env python3
"""Verify cleaned programme Excel workbooks."""

from __future__ import annotations

import re
import argparse
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError as exc:  # pragma: no cover - dependency availability is environment-specific.
    raise SystemExit("openpyxl is required: run with the Codex bundled Python runtime or install openpyxl.") from exc


FINAL_HEADERS = [
    "Institution",
    "Programme",
    "Award",
    "Type / Delivery / Mode",
    "Course and Training Content",
    "Academic Requirements and Restrictions",
    "Application and Research Materials",
    "Application Timing and Status",
    "Fees, Funding, and Special Notes",
    "Official Source",
    "Accessed Date",
]

DELETED = {
    "国家/地区",
    "排名组",
    "排名/学校范围",
    "标准方向组",
    "开设单位/院系",
    "申请状态/可行性",
    "零基础/跨申风险",
    "跨申补证据",
    "适合申请者",
    "Source_sections_checked",
    "Verification_status",
    "Missing_fields",
    "Next_action",
    "Program介绍 + 课程/训练",
    "申请要求",
    "学制/学习方式",
    "费用/资金/重要信息",
    "课程/训练/毕业要求",
}

BANNED = [
    "开设院校为",
    "开设单位为",
    "官网信息显示",
    "项目页或源表记录",
    "官网或源表记录",
    "源表未保留",
    "官方来源列保留",
    "未添加源表外内容",
    "以官网为准",
    "需官网复核",
    "申请者应",
    "建议",
    "适合",
    "风险",
    "可行性",
    "匹配",
    "优先",
    "不建议",
    "可作为",
    "冲刺",
    "保底",
    "竞争激烈",
    "本表",
    "相关性依据",
    "归入",
    "源表未提供",
    "源表未单独列明",
    "源表",
    "按项目页确认",
    "按官方项目页确认",
    "按官方页确认",
    "通俗理解",
    "重点核对",
    "关键不是",
    "申请前",
    "提前申请",
    "按课程页",
    "通常",
    "一般",
    "重点不是",
    "更常",
    "共同评估",
    "判断",
    "为准",
    "Duration/mode uncertain",
    "uncertain from gathered evidence",
    "需核对",
    "需核验",
    "核验",
    "工程/交叉",
]

TYPE_VALUES = {
    "Taught",
    "Research",
    "Taught+Research",
    "Professional",
    "Executive/Professional",
    "Conversion",
    "Directory/Listing",
    "Not stated on the official source",
    "官网未列明",
}

DELIVERY_VALUES = {
    "On-campus",
    "Online",
    "Hybrid",
    "Distance",
    "Block/Residential",
    "Clinical/Placement-based",
    "Field-based",
    "Not stated on the official source",
    "官网未列明",
}

MODE_VALUES = {"FT", "PT", "FT/PT", "Flexible", "Not stated on the official source", "官网未列明"}
COURSE_LABELS = ["Knowledge topics:", "Methods and tools:", "Practical training:", "Programme outputs:"]
ACADEMIC_LABELS = ["Degree and grades:", "Subject background:", "Prerequisites and skills:", "Language:", "Standardised tests:", "Experience, qualifications, and restrictions:"]


def field_value(text: str, name: str) -> str:
    match = re.search(rf"{re.escape(name)}=([^；;。]+)", text)
    return match.group(1).strip() if match else ""


def fail(message: str, failures: list[str]) -> None:
    failures.append(message)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Verify cleaned programme Excel workbooks.")
    parser.add_argument("--dir", required=True, type=Path, help="Directory containing cleaned .xlsx workbooks.")
    parser.add_argument("--files", nargs="*", help="Optional workbook filenames. Defaults to every .xlsx in --dir.")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    workbook_dir = args.dir.expanduser().resolve()
    filenames = args.files or sorted(p.name for p in workbook_dir.glob("*.xlsx") if not p.name.startswith("~$"))
    if not filenames:
        raise SystemExit(f"No .xlsx files found in {workbook_dir}")
    failures: list[str] = []
    sheet_count = 0
    row_count = 0
    for filename in filenames:
        wb = load_workbook(workbook_dir / filename, data_only=True)
        for ws in wb.worksheets:
            sheet_count += 1
            headers = [ws.cell(3, c).value for c in range(1, ws.max_column + 1)]
            if headers != FINAL_HEADERS:
                fail(f"{filename} {ws.title}: bad headers {headers}", failures)
                continue
            if DELETED & set(headers):
                fail(f"{filename} {ws.title}: deleted headers remain {sorted(DELETED & set(headers))}", failures)

            header_to_col = {h: i + 1 for i, h in enumerate(headers)}
            for r in range(4, ws.max_row + 1):
                row_count += 1
                row = {h: str(ws.cell(r, c).value or "") for h, c in header_to_col.items()}
                for h, text in row.items():
                    bad = [word for word in BANNED if word in text]
                    if bad:
                        fail(f"{filename} {ws.title} R{r} {h}: banned {bad} :: {text[:180]}", failures)
                    if re.search(r"(^|[。；;]\s*)\d+[\.、)]\s*", text):
                        fail(f"{filename} {ws.title} R{r} {h}: numbered list :: {text[:180]}", failures)
                    if "\n" in text and h != "Official Source":
                        fail(f"{filename} {ws.title} R{r} {h}: newline", failures)
                if not row["Official Source"].startswith("http"):
                    fail(f"{filename} {ws.title} R{r}: missing official URL", failures)
                if not re.fullmatch(r"\d{4}-\d{2}-\d{2}", row["Accessed Date"]):
                    fail(f"{filename} {ws.title} R{r}: bad access date {row['Accessed Date']}", failures)
                project = row["Type / Delivery / Mode"]
                if not all(part in project for part in ["Type=", "Delivery=", "Mode=", "Duration="]):
                    fail(f"{filename} {ws.title} R{r}: incomplete project type structure :: {project[:180]}", failures)
                if field_value(project, "Type") not in TYPE_VALUES:
                    fail(f"{filename} {ws.title} R{r}: bad Type value :: {project[:180]}", failures)
                if field_value(project, "Delivery") not in DELIVERY_VALUES:
                    fail(f"{filename} {ws.title} R{r}: bad Delivery value :: {project[:180]}", failures)
                if field_value(project, "Mode") not in MODE_VALUES:
                    fail(f"{filename} {ws.title} R{r}: bad Mode value :: {project[:180]}", failures)
                if not field_value(project, "Duration"):
                    fail(f"{filename} {ws.title} R{r}: missing Duration value :: {project[:180]}", failures)
                if re.search(
                    r"deadline|tuition|fee|application|status|opened|closed|截止|开放|关闭|当前申请状态|申请状态|学费|费用|申请材料|材料科学|化工|生物|商科|法律|医学|\bAI\b|人工智能|biology|chemistry|business|law|medicine",
                    project,
                    re.I,
                ):
                    fail(f"{filename} {ws.title} R{r}: mixed project type field :: {project[:180]}", failures)

                course = row["Course and Training Content"]
                if course.strip(".。 ") == "Not stated on the official source":
                    fail(f"{filename} {ws.title} R{r}: course field only says missing", failures)
                if len(course) < 60 and "detailed syllabus" not in course:
                    fail(f"{filename} {ws.title} R{r}: course field too thin :: {course[:180]}", failures)
                if sum(1 for label in COURSE_LABELS if label in course) < 2:
                    fail(f"{filename} {ws.title} R{r}: course field missing structured labels :: {course[:180]}", failures)
                if any(token and token in course for token in [row["Institution"], row["Programme"], row["Award"] if row["Award"] else "__"]):
                    fail(f"{filename} {ws.title} R{r}: duplicate identifier in course field", failures)
                if re.search(r"tuition|学费|deadline|截止|学制|时长|学习方式|开始时间|申请状态|IELTS|TOEFL|PTE|GPA|语言要求", course, re.I):
                    fail(f"{filename} {ws.title} R{r}: mixed course field", failures)

                academic = row["Academic Requirements and Restrictions"]
                if academic.strip(".。 ") == "Not stated on the official source":
                    fail(f"{filename} {ws.title} R{r}: academic field only says missing", failures)
                missing_labels = [label for label in ACADEMIC_LABELS if label not in academic]
                if missing_labels:
                    fail(f"{filename} {ws.title} R{r}: academic field missing labels {missing_labels} :: {academic[:180]}", failures)
                if re.search(r"English Level 执行|Cambridge 常见|视情况|建议|适合|风险|保底|冲刺|匹配", academic):
                    fail(f"{filename} {ws.title} R{r}: subjective or vague academic field :: {academic[:180]}", failures)

            if ws.freeze_panes != "A4":
                fail(f"{filename} {ws.title}: freeze panes not A4", failures)
            if not ws.auto_filter.ref:
                fail(f"{filename} {ws.title}: missing filter", failures)

    if failures:
        print(f"FAIL {len(failures)}")
        for item in failures[:160]:
            print(item)
        raise SystemExit(1)
    print(f"PASS {len(filenames)} files, {sheet_count} sheets, {row_count} data rows")


if __name__ == "__main__":
    main()
