#!/usr/bin/env python3
"""Clean programme Excel workbooks into the official 11-column export format."""

from __future__ import annotations

import argparse
import re
import shutil
from datetime import date
from pathlib import Path

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.cell_range import MultiCellRange
except ImportError as exc:  # pragma: no cover - dependency availability is environment-specific.
    raise SystemExit("openpyxl is required: run with the Codex bundled Python runtime or install openpyxl.") from exc


FINAL_HEADERS = [
    "Institution",
    "Programme",
    "Award",
    "Type / Delivery / Mode",
    "Course and Training Content",
    "Academic Requirements and Restrictions",
    "Application and Research Materials",
    "Application Timing and Status",
    "Fees, Funding, and Special Notes",
    "Official Source",
    "Accessed Date",
]

MISSING = "Not stated on the official source"

BANNED_FRAGMENTS = [
    "开设院校为",
    "开设单位为",
    "官网信息显示",
    "项目页或源表记录",
    "官网或源表记录",
    "源表未保留",
    "官方来源列保留",
    "未添加源表外内容",
    "以官网为准",
    "需官网复核",
    "申请者应",
    "建议",
    "适合",
    "风险",
    "可行性",
    "匹配",
    "优先",
    "不建议",
    "可作为",
    "冲刺",
    "保底",
    "竞争激烈",
    "本表",
    "相关性依据",
    "归入",
    "源表未提供",
    "源表未单独列明",
    "源表",
    "按项目页确认",
    "按官方项目页确认",
    "按官方页确认",
    "通俗理解",
    "重点核对",
    "关键不是",
    "申请前",
    "提前申请",
    "按课程页",
    "通常",
    "一般",
    "重点不是",
    "更常",
    "共同评估",
    "判断",
    "为准",
    "Duration/mode uncertain",
    "uncertain from gathered evidence",
    "需核对",
    "需核验",
    "核验",
    "工程/交叉",
]

PROCESS_LABELS = [
    "项目定位",
    "课程/训练结构",
    "选课/大纲相关重点",
    "官网定位",
    "学术门槛",
    "背景相关",
    "材料清单",
    "现有官方要求摘录",
    "官方要求",
    "学历/成绩",
    "专业背景",
    "英语",
    "研究/材料",
    "其他",
    "原始官方摘录/备注",
    "申请时间/状态信息为",
    "费用、资金或重要信息为",
    "时间和费用",
    "核验边界",
    "申请时间/状态",
    "费用/资金",
    "费用",
    "学费",
    "资金",
    "重要信息",
    "时长/模式",
    "学制/学习方式",
    "学习形态",
]

ACADEMIC_KEYWORDS = [
    "Good",
    "Honours",
    "2:1",
    "First",
    "GPA",
    "degree",
    "学位",
    "成绩",
    "学历",
    "本科",
    "background",
    "背景",
    "prerequisite",
    "先修",
    "IELTS",
    "TOEFL",
    "PTE",
    "GRE",
    "GMAT",
    "English",
    "英语",
    "语言",
    "ATAS",
    "work experience",
    "工作经验",
    "licence",
    "license",
    "registration",
    "资格",
    "external",
    "internal",
    "国际",
]

MATERIAL_KEYWORDS = [
    "CV",
    "resume",
    "résumé",
    "SOP",
    "statement",
    "personal statement",
    "purpose",
    "推荐",
    "reference",
    "recommendation",
    "transcript",
    "成绩单",
    "portfolio",
    "作品",
    "writing sample",
    "sample",
    "research proposal",
    "proposal",
    "supervisor",
    "导师",
    "interview",
    "面试",
    "essay",
    "application portal",
    "supporting document",
    "certificate",
]

TYPE_VALUES = {
    "Taught",
    "Research",
    "Taught+Research",
    "Professional",
    "Executive/Professional",
    "Conversion",
    "Directory/Listing",
}

COURSE_CONTENT_KEYS = [
    "Course and Training Content",
    "课程/训练内容",
    "课程/训练/毕业要求",
    "Program介绍 + 课程/训练",
]

DEGREE_KEYWORDS = [
    "Bachelor",
    "Master",
    "degree",
    "Honours",
    "2:1",
    "2.1",
    "First",
    "upper second",
    "lower second",
    "GPA",
    "学位",
    "学历",
    "本科",
    "硕士",
    "成绩",
]

SUBJECT_KEYWORDS = [
    "subject",
    "discipline",
    "background",
    "major",
    "relevant",
    "related",
    "专业",
    "背景",
    "学科",
    "相关",
]

SKILL_KEYWORDS = [
    "prerequisite",
    "programming",
    "mathematics",
    "statistics",
    "calculus",
    "linear algebra",
    "skill",
    "先修",
    "技能",
    "数学",
    "统计",
    "编程",
]

LANGUAGE_KEYWORDS = [
    "IELTS",
    "TOEFL",
    "PTE",
    "Duolingo",
    "English",
    "英语",
    "语言",
]

TEST_KEYWORDS = ["GRE", "GMAT"]

RESTRICTION_KEYWORDS = [
    "work experience",
    "professional",
    "registration",
    "licence",
    "license",
    "supervisor",
    "internal",
    "external",
    "ATAS",
    "visa",
    "工作经验",
    "资格",
    "注册",
    "限制",
    "导师",
]


def clean(value) -> str:
    text = str(value or "").strip()
    text = text.replace("\r", "\n")
    text = re.sub(r"\s*\n+\s*", "。", text)
    text = re.sub(r"\s+", " ", text)
    text = text.replace("。。", "。")
    return text.strip(" ，,；;。")


def remove_process_phrases(text: str) -> str:
    text = clean(text)
    text = re.sub(r"^.*?为\s+[^。；;]{1,80}项目，开设院校为\s+[^。；;]+。?", "", text)
    text = re.sub(r"^.*?为\s+[^。；;]{1,80}项目。?", "", text)
    for label in PROCESS_LABELS:
        text = re.sub(rf"(^|[。；;]\s*){re.escape(label)}[:：]\s*", r"\1", text)
    replacements = {
        "官网信息显示，": "",
        "项目页或源表记录的课程与训练内容包括：": "",
        "官网或源表记录的申请要求包括：": "",
        "官方来源列保留项目页，未添加源表外内容": "",
        "源表未保留该项目的具体申请要求摘要": "官网未列明",
        "当前周期/截止日需官网复核": "官网未列明",
        "需官网复核": "官网未列明",
        "以项目页为准": "",
        "以官网为准": "",
        "按课程页的 current cycle deadline 和 funding deadline 提前申请": "",
        "按课程页": "",
        "提前申请": "",
        "申请者应": "",
        "建议": "",
        "适合": "",
        "风险": "",
        "可行性": "",
        "匹配": "相关",
        "优先": "",
        "不建议": "",
    }
    for old, new in replacements.items():
        text = text.replace(old, new)
    text = re.sub(r"(^|[。；;]\s*)\d+[\.、)]\s*", r"\1", text)
    text = re.sub(r"。{2,}", "。", text)
    text = text.replace("：。", "：").strip(" ，,；;。")
    return text


def split_units(text: str) -> list[str]:
    text = remove_process_phrases(text)
    units = re.split(r"(?<=[。！？])\s*|[；;]\s*", text)
    out: list[str] = []
    for unit in units:
        unit = remove_process_phrases(unit)
        if not unit or unit in {"官网未列明", "Official programme page"}:
            continue
        if any(bad in unit for bad in BANNED_FRAGMENTS):
            continue
        out.append(unit)
    return out


def join_units(units: list[str], fallback: str, limit: int) -> str:
    seen = set()
    kept = []
    for unit in units:
        key = re.sub(r"\s+", "", unit)
        if not key or key in seen:
            continue
        seen.add(key)
        kept.append(unit)
    text = ". ".join(kept).strip(".。 ")
    if not text:
        text = fallback
    if not text.endswith((".", "。")):
        text += "."
    if len(text) <= limit:
        return text
    cut = text[:limit]
    pos = max(cut.rfind("."), cut.rfind("。"))
    if pos > max(40, limit * 0.55):
        return cut[: pos + 1]
    return cut.rstrip("，,；;。 ") + "."


def official_missing() -> str:
    return f"{MISSING}."


def clean_award(value) -> str:
    text = clean(value)
    text = re.sub(r"[；;，,]?\s*具体以.*$", "", text)
    return text.strip("；;，,。 ")


def first_present(row: dict, keys: list[str]) -> str:
    for key in keys:
        value = clean(row.get(key))
        if value:
            return value
    return ""


def combined_text(row: dict, keys: list[str]) -> str:
    values = []
    seen = set()
    for key in keys:
        value = clean(row.get(key))
        if not value:
            continue
        dedupe_key = re.sub(r"\s+", "", value)
        if dedupe_key in seen:
            continue
        seen.add(dedupe_key)
        values.append(value)
    return "。".join(values)


def contains_any(text: str, keywords: list[str]) -> bool:
    low = text.lower()
    return any(keyword.lower() in low for keyword in keywords)


def contains_test_keyword(text: str) -> bool:
    return bool(re.search(r"\b(?:GRE|GMAT)\b", text, re.I))


def strip_known_label(unit: str) -> str:
    labels = [
        "Knowledge topics",
        "Methods and tools",
        "Practical training",
        "Programme outputs",
        "Degree and grades",
        "Subject background",
        "Prerequisites and skills",
        "Language",
        "Standardised tests",
        "Experience, qualifications, and restrictions",
        "知识主题",
        "方法/工具",
        "实践训练",
        "项目输出",
        "学位/成绩",
        "专业背景",
        "先修/技能",
        "语言",
        "标化",
        "工作/资格/限制",
    ]
    for label in labels:
        unit = re.sub(rf"^{re.escape(label)}[:：]\s*", "", unit)
    return unit.strip(" ，,；;。")


def dedupe_units(units: list[str], limit: int = 3) -> list[str]:
    kept = []
    seen = set()
    for unit in units:
        unit = strip_known_label(remove_process_phrases(unit))
        if not unit or unit == "官网未列明":
            continue
        if any(bad in unit for bad in BANNED_FRAGMENTS):
            continue
        key = re.sub(r"\s+", "", unit).lower()
        if key in seen:
            continue
        seen.add(key)
        kept.append(unit)
        if len(kept) >= limit:
            break
    return kept


def label_value(units: list[str], fallback: str = MISSING, limit: int = 3) -> str:
    kept = dedupe_units(units, limit)
    return "; ".join(kept) if kept else fallback


def programme_blob(row: dict) -> str:
    return combined_text(
        row,
        [
            "Programme",
            "Program",
            "Award",
            "项目类型",
            "项目类型/学习方式",
            "学制/学习方式",
            "申请要求",
            *COURSE_CONTENT_KEYS,
        ],
    )


def explicit_field(text: str, field: str) -> str:
    match = re.search(rf"{re.escape(field)}\s*=\s*([^；;。]+)", text, flags=re.I)
    return clean(match.group(1)) if match else ""


def determine_type(row: dict) -> str:
    text = programme_blob(row)
    explicit = explicit_field(text, "Type")
    if explicit in TYPE_VALUES or explicit in {MISSING, "官网未列明"}:
        return MISSING if explicit == "官网未列明" else explicit
    if re.search(r"directory|listing|candidate list|候选清单|项目目录|列表", text, re.I):
        return "Directory/Listing"
    if re.search(r"executive|高级管理|在职高管", text, re.I):
        return "Executive/Professional"
    if re.search(r"conversion|graduate entry|non-specialist|non-specialists|转换", text, re.I):
        return "Conversion"
    research = re.search(
        r"\b(PhD|DPhil|MPhil|MRes|MSc by Research|MA by Research|by research|by thesis|research degree|supervised research|thesis)\b|研究型",
        text,
        re.I,
    )
    taught = re.search(r"\b(taught|coursework|modules?|lectures?|seminars?)\b|授课型|课程型", text, re.I)
    if research and taught and re.search(r"taught.*research|research.*taught|coursework.*thesis|modules?.*dissertation", text, re.I):
        return "Taught+Research"
    if research:
        return "Research"
    if re.search(r"\b(MBA|MPA|MPH|MEd|JD)\b|professional|practice-based|职业|专业实践", text, re.I):
        return "Professional"
    if taught:
        return "Taught"
    return MISSING


def determine_delivery(row: dict) -> str:
    text = programme_blob(row)
    explicit = explicit_field(text, "Delivery")
    allowed = {
        "On-campus",
        "Online",
        "Hybrid",
        "Distance",
        "Block/Residential",
        "Clinical/Placement-based",
        "Field-based",
        MISSING,
        "官网未列明",
    }
    if explicit in allowed:
        return MISSING if explicit == "官网未列明" else explicit
    if re.search(r"\bonline\b|线上", text, re.I):
        return "Online"
    if re.search(r"\bdistance\b|remote|远程", text, re.I):
        return "Distance"
    if re.search(r"\bhybrid\b|blended|混合", text, re.I):
        return "Hybrid"
    if re.search(r"block teaching|residential|intensive block|集中授课|住宿", text, re.I):
        return "Block/Residential"
    if re.search(r"clinical placement|placement-based|clinic|医院实习|临床", text, re.I):
        return "Clinical/Placement-based"
    if re.search(r"fieldwork|field course|field-based|野外", text, re.I):
        return "Field-based"
    if re.search(r"on[- ]campus|campus-based|in person|face-to-face|线下授课|校内", text, re.I):
        return "On-campus"
    return MISSING


def determine_mode(row: dict) -> str:
    text = programme_blob(row)
    explicit = explicit_field(text, "Mode")
    if explicit in {"FT", "PT", "FT/PT", "Flexible", MISSING, "官网未列明"}:
        return MISSING if explicit == "官网未列明" else explicit
    if re.search(r"flexible|self-paced|弹性", text, re.I):
        return "Flexible"
    ft = re.search(r"full[- ]time|\bFT\b|全日制", text, re.I)
    pt = re.search(r"part[- ]time|\bPT\b|非全日制|兼读", text, re.I)
    if ft and pt:
        return "FT/PT"
    if ft:
        return "FT"
    if pt:
        return "PT"
    return MISSING


def normalize_duration(row: dict) -> str:
    text = programme_blob(row)
    explicit = explicit_field(text, "Duration")
    if explicit:
        return MISSING if explicit == "官网未列明" else explicit

    duration_re = r"((?:\d+(?:\.\d+)?|one|two|three|four|five)\s*(?:months?|years?|个月|年))"
    ft_patterns = [
        rf"{duration_re}[^。；;\n/\d]{{0,35}}?(?:full[- ]time|\bFT\b|全日制)",
        rf"(?:full[- ]time|\bFT\b|全日制)[^。；;\n/]{{0,35}}?{duration_re}",
    ]
    pt_patterns = [
        rf"{duration_re}[^。；;\n/\d]{{0,35}}?(?:part[- ]time|\bPT\b|非全日制|兼读)",
        rf"(?:part[- ]time|\bPT\b|非全日制|兼读)[^。；;\n/]{{0,35}}?{duration_re}",
    ]

    def first_match(patterns: list[str]) -> str:
        for pattern in patterns:
            match = re.search(pattern, text, re.I)
            if match:
                return clean(match.group(1))
        return ""

    ft = first_match(ft_patterns)
    pt = first_match(pt_patterns)
    if ft and pt and ft != pt:
        return f"FT {ft} / PT {pt}"
    if ft:
        return ft
    if pt:
        return pt

    generic = re.search(duration_re, text, re.I)
    if generic:
        return clean(generic.group(1))
    return MISSING


def credits_piece(row: dict) -> str:
    text = programme_blob(row)
    explicit = explicit_field(text, "Credits")
    if explicit and explicit not in {MISSING, "官网未列明"}:
        return explicit
    match = re.search(r"\b\d{1,3}\s*(?:credits?|ECTS|CATS|units?)\b|学分[:：]?\s*\d{1,3}|\d{1,3}\s*学分", text, re.I)
    return clean(match.group(0)) if match else ""


def project_type(row: dict) -> str:
    parts = [
        f"Type={determine_type(row)}",
        f"Delivery={determine_delivery(row)}",
        f"Mode={determine_mode(row)}",
        f"Duration={normalize_duration(row)}",
    ]
    credits = credits_piece(row)
    if credits:
        parts.append(f"Credits={credits}")
    return "; ".join(parts) + "."


def course_training(row: dict) -> str:
    program = first_present(row, ["Programme", "Program"])
    school = first_present(row, ["Institution", "学校"])
    award = clean_award(row.get("Award"))
    categories = {
        "Knowledge topics": [],
        "Methods and tools": [],
        "Practical training": [],
        "Programme outputs": [],
    }
    combined = combined_text(row, COURSE_CONTENT_KEYS)
    for unit in split_units(combined):
        unit = strip_known_label(unit)
        if re.fullmatch(
            r"(taught|research|professional|conversion|online|hybrid|on[- ]campus|campus[- ]based|full[- ]time|part[- ]time)",
            unit,
            re.I,
        ):
            continue
        if re.fullmatch(r"\d+(?:/\d+)?\s*(?:calendar|academic)?\s*years?|\d+(?:/\d+)?\s*months?", unit, re.I):
            continue
        if re.fullmatch(r"(?:FT|PT)\s*\d+(?:[-–]\d+(?:\.\d+)?)?\s*(?:年|years?|months?)", unit, re.I):
            continue
        if re.fullmatch(r"\d+\s*(?:AU|credits?|ECTS|CATS|units?)", unit, re.I):
            continue
        if re.fullmatch(r"(?:fall|spring|summer|autumn|winter)\s+term", unit, re.I):
            continue
        if school and school in unit:
            continue
        if program and program in unit:
            continue
        if award and award in unit:
            continue
        if re.search(
            r"\b\d{1,2}\s*(?:calendar|academic)?\s*(months?|years?)\b|full-time|part-time|deadline|tuition|fee|IELTS|TOEFL|PTE|GPA|学制|时长|学习方式|开始时间|截止|当前申请状态|申请状态|入学|开放|关闭|学费|费用|语言|申请材料",
            unit,
            re.I,
        ):
            continue
        if re.search(r"dissertation|thesis|capstone|portfolio|viva|presentation|report|project output|论文|毕业设计|作品集|答辩|报告", unit, re.I):
            categories["Programme outputs"].append(unit)
        elif re.search(r"method|technique|tool|software|programming|statistics|laboratory method|analysis|实验技术|方法|工具|软件|编程|统计|分析", unit, re.I):
            categories["Methods and tools"].append(unit)
        elif re.search(r"lab|laboratory|fieldwork|placement|internship|workshop|seminar|practical|training|clinic|实践|实验|实习|训练|研讨", unit, re.I):
            categories["Practical training"].append(unit)
        else:
            categories["Knowledge topics"].append(unit)

    knowledge = label_value(categories["Knowledge topics"], fallback=f"{MISSING}; detailed syllabus unavailable")
    methods = label_value(categories["Methods and tools"])
    practice = label_value(categories["Practical training"])
    outputs = label_value(categories["Programme outputs"])
    result = f"Knowledge topics: {knowledge}. Methods and tools: {methods}. Practical training: {practice}. Programme outputs: {outputs}."
    if len(result) < 60 and "detailed syllabus" not in result:
        knowledge = f"{knowledge}; detailed syllabus unavailable"
        result = f"Knowledge topics: {knowledge}. Methods and tools: {methods}. Practical training: {practice}. Programme outputs: {outputs}."
    return result


def select_units_by_keywords(text: str, keywords: list[str]) -> list[str]:
    units = []
    for unit in split_units(text):
        if any(keyword.lower() in unit.lower() for keyword in keywords):
            units.append(unit)
    return units


def academic_requirements(row: dict) -> str:
    categories = {
        "Degree and grades": [],
        "Subject background": [],
        "Prerequisites and skills": [],
        "Language": [],
        "Standardised tests": [],
        "Experience, qualifications, and restrictions": [],
    }
    combined = combined_text(row, ["Academic Requirements and Restrictions", "学术背景/限制条件", "Application Requirements", "申请要求"])
    for unit in split_units(combined):
        unit = strip_known_label(unit)
        academic_signal = (
            contains_any(unit, DEGREE_KEYWORDS)
            or contains_any(unit, SUBJECT_KEYWORDS)
            or contains_any(unit, SKILL_KEYWORDS)
            or contains_any(unit, LANGUAGE_KEYWORDS)
            or contains_test_keyword(unit)
            or contains_any(unit, RESTRICTION_KEYWORDS)
        )
        if contains_any(unit, MATERIAL_KEYWORDS) and not academic_signal:
            continue
        if contains_any(unit, DEGREE_KEYWORDS):
            categories["Degree and grades"].append(unit)
        if contains_any(unit, SUBJECT_KEYWORDS):
            categories["Subject background"].append(unit)
        if contains_any(unit, SKILL_KEYWORDS):
            categories["Prerequisites and skills"].append(unit)
        if contains_any(unit, LANGUAGE_KEYWORDS):
            categories["Language"].append(unit)
        if contains_test_keyword(unit):
            categories["Standardised tests"].append(unit)
        if contains_any(unit, RESTRICTION_KEYWORDS):
            categories["Experience, qualifications, and restrictions"].append(unit)

    standard_tests = label_value(categories["Standardised tests"], fallback=f"GRE/GMAT {MISSING.lower()}", limit=2)
    return (
        f"Degree and grades: {label_value(categories['Degree and grades'], limit=2)}. "
        f"Subject background: {label_value(categories['Subject background'], limit=2)}. "
        f"Prerequisites and skills: {label_value(categories['Prerequisites and skills'], limit=2)}. "
        f"Language: {label_value(categories['Language'], limit=2)}. "
        f"Standardised tests: {standard_tests}. "
        f"Experience, qualifications, and restrictions: {label_value(categories['Experience, qualifications, and restrictions'], limit=2)}."
    )


def material_requirements(row: dict) -> str:
    units = select_units_by_keywords(combined_text(row, ["Application and Research Materials", "申请材料/研究要求", "Application Requirements", "申请要求"]), MATERIAL_KEYWORDS)
    return join_units(units, official_missing(), 200)


def timing_status(row: dict) -> str:
    units = split_units(first_present(row, ["Application Timing and Status", "申请时间/状态"]))
    if not units:
        units = [unit for unit in split_units(first_present(row, ["Application Requirements", "申请要求"])) if re.search(r"deadline|round|open|closed|start|cycle|application|截止|开放|关闭|轮次|入学|开学", unit, re.I)]
    return join_units(units, official_missing(), 300)


def fee_special(row: dict) -> str:
    units = split_units(combined_text(row, ["Fees, Funding, and Special Notes", "费用/资金/特殊事项", "费用/资金/重要信息"]))
    if not units:
        units = [unit for unit in split_units(first_present(row, ["Application Requirements", "申请要求"])) if re.search(r"tuition|fee|funding|scholarship|deposit|ATAS|visa|insurance|学费|费用|资金|奖学金|押金|签证", unit, re.I)]
    return join_units(units, official_missing(), 220)


def source_urls(row: dict) -> str:
    raw = first_present(row, ["Official Source", "官方来源"])
    urls = re.findall(r"https?://[^\s;；,，]+", raw)
    if not urls and raw.startswith("http"):
        urls = [raw]
    return "; ".join(dict.fromkeys(urls)) if urls else raw


def checked_date(row: dict) -> str:
    raw = first_present(row, ["Accessed Date", "核对日期"])
    if re.fullmatch(r"\d{4}-\d{2}-\d{2}", raw):
        return raw
    return date.today().isoformat()


def build_row(row: dict) -> list[str]:
    return [
        first_present(row, ["Institution", "学校"]),
        first_present(row, ["Programme", "Program"]),
        clean_award(row.get("Award")),
        project_type(row),
        course_training(row),
        academic_requirements(row),
        material_requirements(row),
        timing_status(row),
        fee_special(row),
        source_urls(row),
        checked_date(row),
    ]


def estimate_height(values: list[str]) -> float:
    max_lines = 1
    for text in values:
        lines = 0
        for line in str(text or "").split("\n"):
            lines += max(1, int(len(line) / 55) + 1)
        max_lines = max(max_lines, lines)
    return min(168, max(60, 18 + max_lines * 14))


def clear_sheet(ws) -> None:
    try:
        for merge_range in list(ws.merged_cells.ranges):
            ws.unmerge_cells(str(merge_range))
    except Exception:
        ws.merged_cells = MultiCellRange()
    if ws.max_row:
        ws.delete_rows(1, ws.max_row)
    if ws.max_column:
        ws.delete_cols(1, ws.max_column)


def write_and_style(ws, title: str, rows: list[list[str]]) -> None:
    clear_sheet(ws)
    max_col = len(FINAL_HEADERS)
    ws.append([title] + [""] * (max_col - 1))
    ws.append(["Processing note: 11-column official-source structure; type, delivery, mode, duration, and credits are normalised; missing information remains explicit."] + [""] * (max_col - 1))
    ws.append(FINAL_HEADERS)
    for row in rows:
        ws.append(row)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max_col)

    title_fill = PatternFill("solid", fgColor="1F2937")
    note_fill = PatternFill("solid", fgColor="E5E7EB")
    header_fill = PatternFill("solid", fgColor="374151")
    even_fill = PatternFill("solid", fgColor="F9FAFB")
    odd_fill = PatternFill("solid", fgColor="FFFFFF")
    thin = Side(style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:{get_column_letter(max_col)}{ws.max_row}"
    ws.row_dimensions[1].height = 26
    ws.row_dimensions[2].height = 34
    ws.row_dimensions[3].height = 40

    ws["A1"].fill = title_fill
    ws["A1"].font = Font(name="Aptos Display", size=15, bold=True, color="FFFFFF")
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws["A2"].fill = note_fill
    ws["A2"].font = Font(name="Aptos", size=10, color="111827")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    widths = {
        "Institution": 26,
        "Programme": 42,
        "Award": 13,
        "Type / Delivery / Mode": 30,
        "Course and Training Content": 52,
        "Academic Requirements and Restrictions": 58,
        "Application and Research Materials": 46,
        "Application Timing and Status": 48,
        "Fees, Funding, and Special Notes": 42,
        "Official Source": 48,
        "Accessed Date": 14,
    }
    for c, header in enumerate(FINAL_HEADERS, 1):
        letter = get_column_letter(c)
        ws.column_dimensions[letter].width = widths[header]
        cell = ws.cell(3, c)
        cell.fill = header_fill
        cell.font = Font(name="Aptos", size=10, bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    text_cols = [4, 5, 6, 7, 8, 9, 10]
    for r in range(4, ws.max_row + 1):
        fill = even_fill if (r - 4) % 2 == 0 else odd_fill
        for c in range(1, max_col + 1):
            cell = ws.cell(r, c)
            cell.fill = fill
            cell.border = border
            cell.font = Font(name="Aptos", size=9, color="111827")
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        ws.row_dimensions[r].height = estimate_height([ws.cell(r, c).value for c in text_cols])


def process(filename: str, source_dir: Path, out_dir: Path) -> Path:
    wb = load_workbook(source_dir / filename)
    for ws in wb.worksheets:
        headers = {ws.cell(3, c).value: c for c in range(1, ws.max_column + 1)}
        rows = []
        for r in range(4, ws.max_row + 1):
            original = {h: ws.cell(r, c).value for h, c in headers.items()}
            rows.append(build_row(original))
        title = f"{filename.replace('.xlsx', '')} - {ws.title}"
        write_and_style(ws, title, rows)

    out_dir.mkdir(parents=True, exist_ok=True)
    out = out_dir / filename
    wb.save(out)
    return out


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Clean programme Excel workbooks into an 11-column objective official-information structure."
    )
    parser.add_argument("--source-dir", required=True, type=Path, help="Directory containing source .xlsx workbooks.")
    parser.add_argument("--out-dir", required=True, type=Path, help="Directory where cleaned .xlsx workbooks will be written.")
    parser.add_argument("--files", nargs="*", help="Optional workbook filenames. Defaults to every .xlsx in --source-dir.")
    parser.add_argument(
        "--copy-back",
        action="store_true",
        help="Copy cleaned files back over same-named source files after writing them to --out-dir.",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    source_dir = args.source_dir.expanduser().resolve()
    out_dir = args.out_dir.expanduser().resolve()
    filenames = args.files or sorted(p.name for p in source_dir.glob("*.xlsx") if not p.name.startswith("~$"))
    if not filenames:
        raise SystemExit(f"No .xlsx files found in {source_dir}")
    for filename in filenames:
        out = process(filename, source_dir, out_dir)
        print(out)
        if args.copy_back:
            shutil.copy2(out, source_dir / filename)


if __name__ == "__main__":
    main()
