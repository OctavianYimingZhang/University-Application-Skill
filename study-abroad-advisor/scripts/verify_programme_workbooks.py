#!/usr/bin/env python3
"""Verify cleaned programme Excel workbooks."""

from __future__ import annotations

import re
import argparse
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError as exc:  # pragma: no cover - dependency availability is environment-specific.
    raise SystemExit("openpyxl is required: run with the Codex bundled Python runtime or install openpyxl.") from exc


FINAL_HEADERS = [
    "学校",
    "Program",
    "Award",
    "项目类型/学习方式",
    "课程/训练/毕业要求",
    "学术背景/限制条件",
    "申请材料/研究要求",
    "申请时间/状态",
    "费用/资金/特殊事项",
    "官方来源",
    "核对日期",
]

DELETED = {
    "国家/地区",
    "排名组",
    "排名/学校范围",
    "标准方向组",
    "开设单位/院系",
    "申请状态/可行性",
    "零基础/跨申风险",
    "跨申补证据",
    "适合申请者",
    "Source_sections_checked",
    "Verification_status",
    "Missing_fields",
    "Next_action",
    "Program介绍 + 课程/训练",
    "申请要求",
    "学制/学习方式",
    "费用/资金/重要信息",
}

BANNED = [
    "开设院校为",
    "开设单位为",
    "官网信息显示",
    "项目页或源表记录",
    "官网或源表记录",
    "源表未保留",
    "官方来源列保留",
    "未添加源表外内容",
    "以官网为准",
    "需官网复核",
    "申请者应",
    "建议",
    "适合",
    "风险",
    "可行性",
    "匹配",
    "优先",
    "不建议",
    "可作为",
    "冲刺",
    "保底",
    "竞争激烈",
    "本表",
    "相关性依据",
    "归入",
    "源表未提供",
    "源表未单独列明",
    "源表",
    "按项目页确认",
    "按官方项目页确认",
    "按官方页确认",
    "通俗理解",
    "重点核对",
    "关键不是",
    "申请前",
    "提前申请",
    "按课程页",
    "通常",
    "一般",
    "重点不是",
    "更常",
    "共同评估",
    "判断",
    "为准",
]


def fail(message: str, failures: list[str]) -> None:
    failures.append(message)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Verify cleaned programme Excel workbooks.")
    parser.add_argument("--dir", required=True, type=Path, help="Directory containing cleaned .xlsx workbooks.")
    parser.add_argument("--files", nargs="*", help="Optional workbook filenames. Defaults to every .xlsx in --dir.")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    workbook_dir = args.dir.expanduser().resolve()
    filenames = args.files or sorted(p.name for p in workbook_dir.glob("*.xlsx") if not p.name.startswith("~$"))
    if not filenames:
        raise SystemExit(f"No .xlsx files found in {workbook_dir}")
    failures: list[str] = []
    sheet_count = 0
    row_count = 0
    for filename in filenames:
        wb = load_workbook(workbook_dir / filename, data_only=True)
        for ws in wb.worksheets:
            sheet_count += 1
            headers = [ws.cell(3, c).value for c in range(1, ws.max_column + 1)]
            if headers != FINAL_HEADERS:
                fail(f"{filename} {ws.title}: bad headers {headers}", failures)
                continue
            if DELETED & set(headers):
                fail(f"{filename} {ws.title}: deleted headers remain {sorted(DELETED & set(headers))}", failures)

            header_to_col = {h: i + 1 for i, h in enumerate(headers)}
            for r in range(4, ws.max_row + 1):
                row_count += 1
                row = {h: str(ws.cell(r, c).value or "") for h, c in header_to_col.items()}
                for h, text in row.items():
                    bad = [word for word in BANNED if word in text]
                    if bad:
                        fail(f"{filename} {ws.title} R{r} {h}: banned {bad} :: {text[:180]}", failures)
                    if re.search(r"(^|[。；;]\s*)\d+[\.、)]\s*", text):
                        fail(f"{filename} {ws.title} R{r} {h}: numbered list :: {text[:180]}", failures)
                    if "\n" in text and h != "官方来源":
                        fail(f"{filename} {ws.title} R{r} {h}: newline", failures)
                if not row["官方来源"].startswith("http"):
                    fail(f"{filename} {ws.title} R{r}: missing official URL", failures)
                if not re.fullmatch(r"\d{4}-\d{2}-\d{2}", row["核对日期"]):
                    fail(f"{filename} {ws.title} R{r}: bad check date {row['核对日期']}", failures)
                if any(token in row["课程/训练/毕业要求"] for token in [row["学校"], row["Program"], row["Award"] if row["Award"] else "__"]):
                    fail(f"{filename} {ws.title} R{r}: duplicate identifier in course field", failures)
                if re.search(r"tuition|学费|deadline|截止|学制|时长|学习方式|开始时间|申请状态", row["课程/训练/毕业要求"], re.I):
                    fail(f"{filename} {ws.title} R{r}: mixed course field", failures)

            if ws.freeze_panes != "A4":
                fail(f"{filename} {ws.title}: freeze panes not A4", failures)
            if not ws.auto_filter.ref:
                fail(f"{filename} {ws.title}: missing filter", failures)

    if failures:
        print(f"FAIL {len(failures)}")
        for item in failures[:160]:
            print(item)
        raise SystemExit(1)
    print(f"PASS {len(filenames)} files, {sheet_count} sheets, {row_count} data rows")


if __name__ == "__main__":
    main()
