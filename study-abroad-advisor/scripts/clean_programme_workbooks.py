#!/usr/bin/env python3
"""Clean programme Excel workbooks into the official 11-column export format."""

from __future__ import annotations

import argparse
import re
import shutil
from datetime import date
from pathlib import Path

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.cell_range import MultiCellRange
except ImportError as exc:  # pragma: no cover - dependency availability is environment-specific.
    raise SystemExit("openpyxl is required: run with the Codex bundled Python runtime or install openpyxl.") from exc


FINAL_HEADERS = [
    "学校",
    "Program",
    "Award",
    "项目类型/学习方式",
    "课程/训练/毕业要求",
    "学术背景/限制条件",
    "申请材料/研究要求",
    "申请时间/状态",
    "费用/资金/特殊事项",
    "官方来源",
    "核对日期",
]

BANNED_FRAGMENTS = [
    "开设院校为",
    "开设单位为",
    "官网信息显示",
    "项目页或源表记录",
    "官网或源表记录",
    "源表未保留",
    "官方来源列保留",
    "未添加源表外内容",
    "以官网为准",
    "需官网复核",
    "申请者应",
    "建议",
    "适合",
    "风险",
    "可行性",
    "匹配",
    "优先",
    "不建议",
    "可作为",
    "冲刺",
    "保底",
    "竞争激烈",
    "本表",
    "相关性依据",
    "归入",
    "源表未提供",
    "源表未单独列明",
    "源表",
    "按项目页确认",
    "按官方项目页确认",
    "按官方页确认",
    "通俗理解",
    "重点核对",
    "关键不是",
    "申请前",
    "提前申请",
    "按课程页",
    "通常",
    "一般",
    "重点不是",
    "更常",
    "共同评估",
    "判断",
    "为准",
]

PROCESS_LABELS = [
    "项目定位",
    "课程/训练结构",
    "选课/大纲相关重点",
    "官网定位",
    "学术门槛",
    "背景相关",
    "材料清单",
    "现有官方要求摘录",
    "官方要求",
    "学历/成绩",
    "专业背景",
    "英语",
    "研究/材料",
    "其他",
    "原始官方摘录/备注",
    "申请时间/状态信息为",
    "费用、资金或重要信息为",
    "时间和费用",
    "核验边界",
    "申请时间/状态",
    "费用/资金",
    "费用",
    "学费",
    "资金",
    "重要信息",
    "时长/模式",
    "学制/学习方式",
    "学习形态",
]

ACADEMIC_KEYWORDS = [
    "Good",
    "Honours",
    "2:1",
    "First",
    "GPA",
    "degree",
    "学位",
    "成绩",
    "学历",
    "本科",
    "background",
    "背景",
    "prerequisite",
    "先修",
    "IELTS",
    "TOEFL",
    "PTE",
    "GRE",
    "GMAT",
    "English",
    "英语",
    "语言",
    "ATAS",
    "work experience",
    "工作经验",
    "licence",
    "license",
    "registration",
    "资格",
    "external",
    "internal",
    "国际",
]

MATERIAL_KEYWORDS = [
    "CV",
    "resume",
    "résumé",
    "SOP",
    "statement",
    "personal statement",
    "purpose",
    "推荐",
    "reference",
    "recommendation",
    "transcript",
    "成绩单",
    "portfolio",
    "作品",
    "writing sample",
    "sample",
    "research proposal",
    "proposal",
    "supervisor",
    "导师",
    "interview",
    "面试",
    "essay",
    "application portal",
    "supporting document",
    "certificate",
]


def clean(value) -> str:
    text = str(value or "").strip()
    text = text.replace("\r", "\n")
    text = re.sub(r"\s*\n+\s*", "。", text)
    text = re.sub(r"\s+", " ", text)
    text = text.replace("。。", "。")
    return text.strip(" ，,；;。")


def remove_process_phrases(text: str) -> str:
    text = clean(text)
    text = re.sub(r"^.*?为\s+[^。；;]{1,80}项目，开设院校为\s+[^。；;]+。?", "", text)
    text = re.sub(r"^.*?为\s+[^。；;]{1,80}项目。?", "", text)
    for label in PROCESS_LABELS:
        text = re.sub(rf"(^|[。；;]\s*){re.escape(label)}[:：]\s*", r"\1", text)
    replacements = {
        "官网信息显示，": "",
        "项目页或源表记录的课程与训练内容包括：": "",
        "官网或源表记录的申请要求包括：": "",
        "官方来源列保留项目页，未添加源表外内容": "",
        "源表未保留该项目的具体申请要求摘要": "官网未列明",
        "当前周期/截止日需官网复核": "官网未列明",
        "需官网复核": "官网未列明",
        "以项目页为准": "",
        "以官网为准": "",
        "按课程页的 current cycle deadline 和 funding deadline 提前申请": "",
        "按课程页": "",
        "提前申请": "",
        "申请者应": "",
        "建议": "",
        "适合": "",
        "风险": "",
        "可行性": "",
        "匹配": "相关",
        "优先": "",
        "不建议": "",
    }
    for old, new in replacements.items():
        text = text.replace(old, new)
    text = re.sub(r"(^|[。；;]\s*)\d+[\.、)]\s*", r"\1", text)
    text = re.sub(r"。{2,}", "。", text)
    text = text.replace("：。", "：").strip(" ，,；;。")
    return text


def split_units(text: str) -> list[str]:
    text = remove_process_phrases(text)
    units = re.split(r"(?<=[。！？])\s*|[；;]\s*", text)
    out: list[str] = []
    for unit in units:
        unit = remove_process_phrases(unit)
        if not unit or unit in {"官网未列明", "Official programme page"}:
            continue
        if any(bad in unit for bad in BANNED_FRAGMENTS):
            continue
        out.append(unit)
    return out


def join_units(units: list[str], fallback: str, limit: int) -> str:
    seen = set()
    kept = []
    for unit in units:
        key = re.sub(r"\s+", "", unit)
        if not key or key in seen:
            continue
        seen.add(key)
        kept.append(unit)
    text = "。".join(kept).strip("。")
    if not text:
        text = fallback
    if not text.endswith("。"):
        text += "。"
    if len(text) <= limit:
        return text
    cut = text[:limit]
    pos = cut.rfind("。")
    if pos > max(40, limit * 0.55):
        return cut[: pos + 1]
    return cut.rstrip("，,；;。") + "。"


def official_missing() -> str:
    return "官网未列明。"


def clean_award(value) -> str:
    text = clean(value)
    text = re.sub(r"[；;，,]?\s*具体以.*$", "", text)
    return text.strip("；;，,。 ")


def normalize_type(value: str) -> str:
    text = clean(value)
    terms = []
    if re.search(r"research|研究", text, re.I):
        terms.append("Research")
    if re.search(r"taught|授课", text, re.I):
        terms.append("Taught")
    if re.search(r"professional|practice|职业|专业", text, re.I):
        terms.append("Professional")
    if re.search(r"conversion|转换", text, re.I):
        terms.append("Conversion")
    if re.search(r"online|线上", text, re.I):
        terms.append("Online")
    if re.search(r"hybrid|混合", text, re.I):
        terms.append("Hybrid")
    if not terms and text and not any(bad in text for bad in BANNED_FRAGMENTS):
        terms.append(text)
    return "；".join(dict.fromkeys(terms))


def duration_piece(value: str) -> str:
    raw = str(value or "").replace("\r", "\n")
    for label in ["时长/模式", "学制/学习方式"]:
        match = re.search(rf"{re.escape(label)}[:：]\s*(.*)", raw, flags=re.S)
        if match:
            raw = match.group(1)
            break
    stop = re.search(r"(?:\n|\s)(申请时间|学习形态|费用|重要信息|项目路径)[:：]", raw)
    if stop:
        raw = raw[: stop.start()]
    text = remove_process_phrases(raw)
    text = re.sub(r"\b(research|taught|professional)\b", "", text, flags=re.I)
    text = text.strip(" ，,；;。")
    if any(bad in text for bad in BANNED_FRAGMENTS):
        return ""
    return text


def project_type(row: dict) -> str:
    parts = [normalize_type(row.get("项目类型", "")), duration_piece(row.get("学制/学习方式", ""))]
    parts = [part for part in parts if part]
    text = "；".join(dict.fromkeys(parts))
    return join_units([text], official_missing(), 80)


def course_training(row: dict) -> str:
    program = clean(row.get("Program"))
    school = clean(row.get("学校"))
    award = clean_award(row.get("Award"))
    units = []
    for unit in split_units(row.get("Program介绍 + 课程/训练", "")):
        if school and school in unit:
            continue
        if program and program in unit:
            continue
        if award and award in unit:
            continue
        if re.search(
            r"\b\d{1,2}\s*(months?|years?)\b|full-time|part-time|deadline|tuition|fee|IELTS|TOEFL|GPA|学制|时长|学习方式|开始时间|截止|当前申请状态|申请状态|入学|开放|关闭",
            unit,
            re.I,
        ):
            continue
        units.append(unit)
    return join_units(units, official_missing(), 220)


def select_units_by_keywords(text: str, keywords: list[str]) -> list[str]:
    units = []
    for unit in split_units(text):
        if any(keyword.lower() in unit.lower() for keyword in keywords):
            units.append(unit)
    return units


def academic_requirements(row: dict) -> str:
    units = []
    for unit in select_units_by_keywords(row.get("申请要求", ""), ACADEMIC_KEYWORDS):
        if any(keyword.lower() in unit.lower() for keyword in MATERIAL_KEYWORDS) and not re.search(
            r"Good|Honours|2:1|First|GPA|degree|IELTS|TOEFL|PTE|GRE|GMAT|ATAS|工作经验|work experience|学位|学历|本科",
            unit,
            re.I,
        ):
            continue
        units.append(unit)
    return join_units(units, official_missing(), 260)


def material_requirements(row: dict) -> str:
    units = select_units_by_keywords(row.get("申请要求", ""), MATERIAL_KEYWORDS)
    return join_units(units, official_missing(), 200)


def timing_status(row: dict) -> str:
    units = split_units(row.get("申请时间/状态", ""))
    if not units:
        units = [unit for unit in split_units(row.get("申请要求", "")) if re.search(r"deadline|round|open|closed|start|cycle|application|截止|开放|关闭|轮次|入学|开学", unit, re.I)]
    return join_units(units, official_missing(), 300)


def fee_special(row: dict) -> str:
    units = split_units(row.get("费用/资金/重要信息", ""))
    if not units:
        units = [unit for unit in split_units(row.get("申请要求", "")) if re.search(r"tuition|fee|funding|scholarship|deposit|ATAS|visa|insurance|学费|费用|资金|奖学金|押金|签证", unit, re.I)]
    return join_units(units, official_missing(), 220)


def source_urls(row: dict) -> str:
    raw = clean(row.get("官方来源"))
    urls = re.findall(r"https?://[^\s;；,，]+", raw)
    if not urls and raw.startswith("http"):
        urls = [raw]
    return "；".join(dict.fromkeys(urls)) if urls else raw


def checked_date(row: dict) -> str:
    raw = clean(row.get("核对日期"))
    if re.fullmatch(r"\d{4}-\d{2}-\d{2}", raw):
        return raw
    return date.today().isoformat()


def build_row(row: dict) -> list[str]:
    return [
        clean(row.get("学校")),
        clean(row.get("Program")),
        clean_award(row.get("Award")),
        project_type(row),
        course_training(row),
        academic_requirements(row),
        material_requirements(row),
        timing_status(row),
        fee_special(row),
        source_urls(row),
        checked_date(row),
    ]


def estimate_height(values: list[str]) -> float:
    max_lines = 1
    for text in values:
        lines = 0
        for line in str(text or "").split("\n"):
            lines += max(1, int(len(line) / 55) + 1)
        max_lines = max(max_lines, lines)
    return min(168, max(60, 18 + max_lines * 14))


def clear_sheet(ws) -> None:
    try:
        for merge_range in list(ws.merged_cells.ranges):
            ws.unmerge_cells(str(merge_range))
    except Exception:
        ws.merged_cells = MultiCellRange()
    if ws.max_row:
        ws.delete_rows(1, ws.max_row)
    if ws.max_column:
        ws.delete_cols(1, ws.max_column)


def write_and_style(ws, title: str, rows: list[list[str]]) -> None:
    clear_sheet(ws)
    max_col = len(FINAL_HEADERS)
    ws.append([title] + [""] * (max_col - 1))
    ws.append(["处理说明：11列客观官网信息结构；删除排名、地区、院系、主观判断和内部QA字段；缺失信息仅短句标明。"] + [""] * (max_col - 1))
    ws.append(FINAL_HEADERS)
    for row in rows:
        ws.append(row)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max_col)

    title_fill = PatternFill("solid", fgColor="1F2937")
    note_fill = PatternFill("solid", fgColor="E5E7EB")
    header_fill = PatternFill("solid", fgColor="374151")
    even_fill = PatternFill("solid", fgColor="F9FAFB")
    odd_fill = PatternFill("solid", fgColor="FFFFFF")
    thin = Side(style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:{get_column_letter(max_col)}{ws.max_row}"
    ws.row_dimensions[1].height = 26
    ws.row_dimensions[2].height = 34
    ws.row_dimensions[3].height = 40

    ws["A1"].fill = title_fill
    ws["A1"].font = Font(name="Aptos Display", size=15, bold=True, color="FFFFFF")
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws["A2"].fill = note_fill
    ws["A2"].font = Font(name="Aptos", size=10, color="111827")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    widths = {
        "学校": 26,
        "Program": 42,
        "Award": 13,
        "项目类型/学习方式": 30,
        "课程/训练/毕业要求": 52,
        "学术背景/限制条件": 58,
        "申请材料/研究要求": 46,
        "申请时间/状态": 48,
        "费用/资金/特殊事项": 42,
        "官方来源": 48,
        "核对日期": 14,
    }
    for c, header in enumerate(FINAL_HEADERS, 1):
        letter = get_column_letter(c)
        ws.column_dimensions[letter].width = widths[header]
        cell = ws.cell(3, c)
        cell.fill = header_fill
        cell.font = Font(name="Aptos", size=10, bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    text_cols = [4, 5, 6, 7, 8, 9, 10]
    for r in range(4, ws.max_row + 1):
        fill = even_fill if (r - 4) % 2 == 0 else odd_fill
        for c in range(1, max_col + 1):
            cell = ws.cell(r, c)
            cell.fill = fill
            cell.border = border
            cell.font = Font(name="Aptos", size=9, color="111827")
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        ws.row_dimensions[r].height = estimate_height([ws.cell(r, c).value for c in text_cols])


def process(filename: str, source_dir: Path, out_dir: Path) -> Path:
    wb = load_workbook(source_dir / filename)
    for ws in wb.worksheets:
        headers = {ws.cell(3, c).value: c for c in range(1, ws.max_column + 1)}
        rows = []
        for r in range(4, ws.max_row + 1):
            original = {h: ws.cell(r, c).value for h, c in headers.items()}
            rows.append(build_row(original))
        title = f"{filename.replace('.xlsx', '')} - {ws.title}"
        write_and_style(ws, title, rows)

    out_dir.mkdir(parents=True, exist_ok=True)
    out = out_dir / filename
    wb.save(out)
    return out


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Clean programme Excel workbooks into an 11-column objective official-information structure."
    )
    parser.add_argument("--source-dir", required=True, type=Path, help="Directory containing source .xlsx workbooks.")
    parser.add_argument("--out-dir", required=True, type=Path, help="Directory where cleaned .xlsx workbooks will be written.")
    parser.add_argument("--files", nargs="*", help="Optional workbook filenames. Defaults to every .xlsx in --source-dir.")
    parser.add_argument(
        "--copy-back",
        action="store_true",
        help="Copy cleaned files back over same-named source files after writing them to --out-dir.",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    source_dir = args.source_dir.expanduser().resolve()
    out_dir = args.out_dir.expanduser().resolve()
    filenames = args.files or sorted(p.name for p in source_dir.glob("*.xlsx") if not p.name.startswith("~$"))
    if not filenames:
        raise SystemExit(f"No .xlsx files found in {source_dir}")
    for filename in filenames:
        out = process(filename, source_dir, out_dir)
        print(out)
        if args.copy_back:
            shutil.copy2(out, source_dir / filename)


if __name__ == "__main__":
    main()
